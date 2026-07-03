# utils/excel_value_loader.py
"""从 Excel 矩阵文件加载信号值描述

支持多种实际格式：
- .xlsx (OpenXML，使用 openpyxl)
- .xls (OLE2 复合文档，使用 xlrd 1.x 或其他引擎)
- HTML 表格保存为 .xls（中国车企工具常见）
- CSV 保存为 .xls

自动检测文件真实格式，按策略依次尝试多种读取方式。
"""

import os
import re
import pandas as pd
import xlrd


# 信号名称列的匹配关键词（按优先级排列）
SIGNAL_NAME_KEYWORDS = [
    "signal name",       # 英文
    "signal_name",
    "信号名",
    "信号名称",
    "signal",            # 更宽泛的匹配
]

# 信号值描述列的匹配关键词（按优先级排列）
VALUE_DESC_KEYWORDS = [
    "signal value description",
    "信号值描述",
    "value description",
    "值描述",
    "value desc",
    "信号描述",          # 备选
    "signal description",
]


def load_value_descriptions(file_path: str) -> dict[str, dict[int, str]]:
    """加载 Excel 值描述文件，自动检测实际格式。

    按以下策略依次尝试读取：
    1. openpyxl (标准 .xlsx)
    2. xlrd 直接读取 OLE2 .xls 文件（绕过 pandas 版本限制）
    3. pandas xlrd引擎（仅适用于 .xlsx，需 xlrd>=2.0.1）
    4. pandas 默认引擎（自动推断）
    5. HTML 表格（很多中国工具导出的"Excel"实际是 HTML）
    6. CSV/TSV（可能是文本文件保存为 .xls）
    7. 无表头自动推断
    8. HTML 无表头推断

    Args:
        file_path: Excel 文件路径（.xlsx / .xls / HTML / CSV）

    Returns:
        {"SignalName": {0: "OFF", 1: "ON"}, ...}

    Raises:
        ValueError: 未找到必要的列或无法读取文件时抛出
    """
    ext = os.path.splitext(file_path)[1].lower()

    df = _try_read_excel(file_path, ext)

    if df is None:
        # 读取文件头部帮助诊断实际格式
        _raise_format_error(file_path)

    # 在成功读取的 DataFrame 中查找目标列
    name_col = _find_column(df, SIGNAL_NAME_KEYWORDS)
    desc_col = _find_column(df, VALUE_DESC_KEYWORDS)

    if name_col is None:
        raise ValueError(
            f"未找到信号名称列。当前列名: {list(df.columns)}\n"
            f"尝试的关键词: {SIGNAL_NAME_KEYWORDS}"
        )
    if desc_col is None:
        raise ValueError(
            f"未找到信号值描述列。当前列名: {list(df.columns)}\n"
            f"尝试的关键词: {VALUE_DESC_KEYWORDS}"
        )

    descriptions: dict[str, dict[int, str]] = {}
    for _, row in df.iterrows():
        sig_name = str(row[name_col]).strip() if pd.notna(row[name_col]) else ""
        desc_text = str(row[desc_col]) if pd.notna(row[desc_col]) else ""

        # 跳过空行或 NaN
        if not sig_name or sig_name == "nan":
            continue

        parsed = parse_value_description(desc_text)
        if parsed:
            if sig_name in descriptions:
                # 同一信号出现多行时合并描述
                descriptions[sig_name].update(parsed)
            else:
                descriptions[sig_name] = parsed

    return descriptions


def _read_xls_via_xlrd_directly(file_path: str, header_row: int = 0) -> pd.DataFrame | None:
    """直接使用 xlrd 读取 OLE2 .xls 文件并转为 DataFrame。

    绕过 pandas 的 xlrd 版本限制（pandas>=3.x 要求 xlrd>=2.0.1，
    但 xlrd 2.x 已移除 .xls OLE2 支持），直接调用 xlrd 1.2.0 API。

    CAN矩阵文件通常有多个sheet（Macro1、Cover、History、Legend、Matrix），
    数据在 "Matrix" sheet中。因此需要遍历所有sheet查找包含目标列的那个。

    Args:
        file_path: .xls 文件路径
        header_row: 用作表头的行索引（0 或 1）

    Returns:
        成功时返回 DataFrame，失败时返回 None
    """
    try:
        wb = xlrd.open_workbook(file_path)
    except Exception:
        return None

    # 遍历所有sheet，找到包含信号名称和值描述列的那个
    for ws in wb.sheets():
        if ws.nrows == 0:
            continue

        # 提取所有行数据
        all_rows = []
        for row_idx in range(ws.nrows):
            row_data = []
            for col_idx in range(ws.ncols):
                cell = ws.cell(row_idx, col_idx)
                # xlrd cell 类型: 0=empty, 1=text, 2=number, 3=date, 4=boolean, 5=error
                if cell.ctype == 0:  # 空单元格
                    row_data.append('')
                elif cell.ctype == 2:  # 数字
                    # 如果是整数则显示为整数，否则显示为浮点数
                    if cell.value == int(cell.value):
                        row_data.append(str(int(cell.value)))
                    else:
                        row_data.append(str(cell.value))
                elif cell.ctype == 3:  # 日期
                    # 将 xlrd 日期元组转为字符串
                    date_tuple = xlrd.xldate_as_tuple(cell.value, wb.datemode)
                    row_data.append(str(date_tuple))
                elif cell.ctype == 4:  # boolean
                    row_data.append(str(bool(cell.value)))
                elif cell.ctype == 5:  # error
                    row_data.append('')
                else:  # text (ctype==1) 及其他
                    row_data.append(str(cell.value))
            all_rows.append(row_data)

        # 尝试不同的表头行位置（CAN矩阵表头可能在row 0~3）
        for hr in range(min(5, len(all_rows))):
            headers = all_rows[hr]
            data_start = hr + 1
            data_rows = all_rows[data_start:]

            if not data_rows:
                continue

            # 确保所有数据行的列数与表头一致
            max_cols = len(headers)
            for i in range(len(data_rows)):
                if len(data_rows[i]) < max_cols:
                    data_rows[i].extend([''] * (max_cols - len(data_rows[i])))
                elif len(data_rows[i]) > max_cols:
                    data_rows[i] = data_rows[i][:max_cols]

            df = pd.DataFrame(data_rows, columns=headers)
            if _has_target_columns(df):
                return df

    return None


def _try_read_excel(file_path: str, ext: str) -> pd.DataFrame | None:
    """尝试多种方式读取 Excel 文件，返回包含目标列的 DataFrame。

    依次尝试各种策略，每种策略内尝试不同的 header 行位置。
    找到包含目标列的 DataFrame 即返回，全部失败则返回 None。
    """
    # 策略1: openpyxl (标准 .xlsx)——遍历所有sheet查找包含目标列的
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            for header_row in [0, 1, 2, 3]:
                try:
                    df = pd.read_excel(
                        file_path, engine='openpyxl',
                        sheet_name=sheet_name, header=header_row
                    )
                    if _has_target_columns(df):
                        wb.close()
                        return df
                except Exception:
                    pass
        wb.close()
    except Exception:
        pass

    # 策略2: xlrd 直接读取 OLE2 .xls 文件（内部已遍历所有sheet和表头行）
    # pandas >=3.x 要求 xlrd>=2.0.1，但 xlrd 2.x 已移除 .xls 支持。
    # 因此对 OLE2 文件必须绕过 pandas，直接用 xlrd 1.2.0 读取再转 DataFrame。
    try:
        df = _read_xls_via_xlrd_directly(file_path)
        if df is not None and _has_target_columns(df):
            return df
    except Exception:
        pass

    # 策略3: pandas xlrd引擎（仅适用于 xlrd 2.x 可读的 .xlsx 文件）
    for header_row in [0, 1]:
        try:
            df = pd.read_excel(file_path, engine='xlrd', header=header_row)
            if _has_target_columns(df):
                return df
        except Exception:
            pass

    # 策略4: pandas 默认引擎（让 pandas 自动推断格式）
    for header_row in [0, 1]:
        try:
            df = pd.read_excel(file_path, header=header_row)
            if _has_target_columns(df):
                return df
        except Exception:
            pass

    # 策略5: HTML 表格（很多中国车企工具导出的是 HTML 保存为 .xls）
    for encoding in ['utf-8', 'gbk', 'gb2312', 'gb18030']:
        try:
            tables = pd.read_html(file_path, encoding=encoding)
            for df in tables:
                if _has_target_columns(df):
                    return df
        except Exception:
            pass

    # 策略6: CSV/TSV（可能是 CSV 保存为 .xls）
    for encoding in ['utf-8', 'gbk', 'gb2312', 'gb18030', 'utf-8-sig']:
        for sep in [',', '\t', ';']:
            try:
                df = pd.read_csv(file_path, encoding=encoding, sep=sep)
                if _has_target_columns(df):
                    return df
            except Exception:
                pass

    # 策略7: 无表头，自动推断——先整体读取，再尝试把各行当表头
    try:
        df_raw = pd.read_excel(file_path, header=None)
        if len(df_raw) > 0:
            for i in range(min(5, len(df_raw))):
                test_df = df_raw.copy()
                test_df.columns = test_df.iloc[i].astype(str)
                test_df = test_df.iloc[i + 1:].reset_index(drop=True)
                if _has_target_columns(test_df):
                    return test_df
    except Exception:
        pass

    # 策略8: HTML 无表头推断（同策略7逻辑，但用 read_html）
    for encoding in ['utf-8', 'gbk', 'gb18030']:
        try:
            tables = pd.read_html(file_path, encoding=encoding, header=None)
            for df_raw in tables:
                if len(df_raw) > 0:
                    for i in range(min(5, len(df_raw))):
                        test_df = df_raw.copy()
                        test_df.columns = test_df.iloc[i].astype(str)
                        test_df = test_df.iloc[i + 1:].reset_index(drop=True)
                        if _has_target_columns(test_df):
                            return test_df
        except Exception:
            pass

    return None


def _has_target_columns(df: pd.DataFrame) -> bool:
    """检查 DataFrame 是否包含目标列（信号名称列 + 值描述列）。"""
    name_col = _find_column(df, SIGNAL_NAME_KEYWORDS)
    desc_col = _find_column(df, VALUE_DESC_KEYWORDS)
    return name_col is not None and desc_col is not None


def _raise_format_error(file_path: str) -> None:
    """读取文件头部字节诊断实际格式，抛出包含诊断信息的 ValueError。"""
    try:
        with open(file_path, 'rb') as f:
            header_bytes = f.read(200)

        # 判断实际格式
        lower_bytes = header_bytes.lower()
        if b'<html' in lower_bytes or b'<table' in lower_bytes or b'<tr' in lower_bytes:
            actual_format = "HTML 表格（伪装为 .xls）"
        elif header_bytes[:2] == b'PK':
            actual_format = "ZIP/OpenXML (.xlsx)"
        elif header_bytes[:8] == b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1':
            actual_format = "OLE2 复合文档 (.xls)"
            # xlrd 直接读取也失败了，说明文件可能损坏或格式异常
            actual_format += " —— xlrd 直接读取失败，请检查文件是否损坏"
        elif header_bytes[:3] == b'\xef\xbb\xbf' or b',' in header_bytes[:100] or b'\t' in header_bytes[:100]:
            actual_format = "CSV/文本文件（伪装为 .xls）"
        else:
            actual_format = f"未知格式 (首字节: {header_bytes[:8].hex()})"

        # 文件头部预览（用于调试）
        header_preview = header_bytes[:80].decode('utf-8', errors='replace')

        raise ValueError(
            f"无法读取文件 {os.path.basename(file_path)}。\n"
            f"检测到的实际格式: {actual_format}\n"
            f"已尝试策略: openpyxl(xlsx), xlrd直接读取, pandas xlrd引擎, 默认引擎, HTML表格, CSV, 无表头推断\n"
            f"请确认文件未损坏且格式正确。\n\n"
            f"文件头部预览: {header_preview}..."
        )
    except ValueError:
        raise
    except Exception as e2:
        raise ValueError(f"无法读取文件: {e2}")


def _find_column(df: pd.DataFrame, keywords: list[str]) -> str | None:
    """在 DataFrame 列名中模糊查找。

    处理双语表头（换行分隔）、多余空格、大小写差异。
    按关键词优先级返回第一个匹配的列名。

    Args:
        df: 数据框
        keywords: 要匹配的关键词列表（按优先级排列，不区分大小写）

    Returns:
        匹配到的原始列名，未找到则返回 None
    """
    # 预处理列名：小写化、替换换行/多余空格
    col_map = {}  # normalized_name -> original_name
    for col in df.columns:
        normalized = str(col).lower()
        normalized = re.sub(r'[\n\r]+', ' ', normalized)  # 换行→空格
        normalized = re.sub(r'\s+', ' ', normalized).strip()  # 多余空格→单空格
        col_map[normalized] = col

    for kw in keywords:
        kw_lower = kw.lower().strip()
        # 精确匹配
        if kw_lower in col_map:
            return col_map[kw_lower]
        # 子串匹配
        for normalized, original in col_map.items():
            if kw_lower in normalized:
                return original

    return None


def parse_value_description(desc_text: str) -> dict[int, str]:
    """解析信号值描述文本。

    支持多种格式：
      - "0:OFF, 1:ON"
      - "0=OFF, 1=ON"
      - "0 - OFF, 1 - ON"
      - "0x0:OFF, 0x1:ON"
      - "0-OFF; 1-ON"
      - 中英文逗号/分号/换行分隔

    Args:
        desc_text: 原始描述文本

    Returns:
        {整数值: "描述字符串", ...}，解析失败返回空字典
    """
    result: dict[int, str] = {}
    if not desc_text or pd.isna(desc_text):
        return result

    desc_text = str(desc_text).strip()
    if not desc_text:
        return result

    # 按逗号、分号、换行等常见分隔符拆分
    pairs = re.split(r'[,;，；\n\r]+', desc_text)

    for pair in pairs:
        pair = pair.strip()
        if not pair:
            continue
        # 匹配 "数值 分隔符 描述" 模式
        # 分隔符支持：冒号、等号、各种横杠（-–—）
        match = re.match(
            r'^(0[xX][\da-fA-F]+|\d+)\s*[:=\-–—]\s*(.+)$', pair
        )
        if match:
            val_str = match.group(1)
            desc = match.group(2).strip()
            try:
                if val_str.startswith(('0x', '0X')):
                    val = int(val_str, 16)
                else:
                    val = int(val_str)
                result[val] = desc
            except ValueError:
                pass

    return result